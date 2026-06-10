"""
UnlistedZone.com Scraper v4
============================
- Yearly financials only (last 3 years max)
- Full description (no truncation, clicks "Read more")
- All fundamentals
- Names from detail page <h1>

pip install playwright beautifulsoup4 openpyxl requests
playwright install chromium

python unlistedzone_scraper_v4.py --limit 5    # test
python unlistedzone_scraper_v4.py               # full run
"""

import argparse, csv, json, re, sys, time
from datetime import datetime
from pathlib import Path
from bs4 import BeautifulSoup

BASE_URL = "https://unlistedzone.com"
DELAY = 2.5
OUTPUT_DIR = Path("output")

FUND_KEYS = [
    "Lot Size", "52 Week High", "52 Week Low", "Depository",
    "PAN Number", "ISIN Number", "CIN", "RTA",
    "Market Cap (in cr.)", "P/E Ratio", "P/B Ratio",
    "Debt to Equity", "ROE (%)", "Book Value", "Face Value", "Total Shares",
]

def cv(v):
    """Clean value string."""
    return v.replace("₹", "").replace(",", "").strip()


# ═══════════════════════════════════════════════════════════════════════
# STEP 1: LISTING PAGE — GET ALL SHARE URLs
# ═══════════════════════════════════════════════════════════════════════
def scrape_listing_page():
    from playwright.sync_api import sync_playwright
    print("[1/3] Loading listing page...")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{BASE_URL}/shares", wait_until="networkidle", timeout=60000)
        time.sleep(3)

        clicks = 0
        while True:
            try:
                btns = page.locator("text=View More")
                clicked = False
                for i in range(btns.count()):
                    btn = btns.nth(i)
                    href = btn.get_attribute("href") or ""
                    if "javascript:void" in href and btn.is_visible(timeout=2000):
                        prev = page.locator("a[href*='/shares/']").count()
                        btn.click()
                        time.sleep(2)
                        page.wait_for_load_state("networkidle", timeout=15000)
                        curr = page.locator("a[href*='/shares/']").count()
                        clicks += 1
                        print(f"   View More #{clicks} — {curr} links")
                        if curr == prev:
                            time.sleep(3)
                            if page.locator("a[href*='/shares/']").count() == prev:
                                clicked = False
                                break
                        clicked = True
                        break
                if not clicked:
                    break
            except Exception:
                break

        html = page.content()
        browser.close()

    soup = BeautifulSoup(html, "html.parser")
    seen, urls = set(), []
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith("/shares/"):
            full = BASE_URL + href
        elif "unlistedzone.com/shares/" in href:
            full = href.replace("https://www.unlistedzone.com", BASE_URL)
        else:
            continue
        path = full.replace(BASE_URL, "").rstrip("/")
        if path in ("/shares", "") or not re.match(r"^/shares/[a-z0-9]", path):
            continue
        clean = full.rstrip("/") + "/"
        if clean not in seen:
            seen.add(clean)
            urls.append(clean)

    print(f"   {len(urls)} unique shares found")
    return urls


# ═══════════════════════════════════════════════════════════════════════
# STEP 2: DETAIL PAGE — FULL EXTRACTION
# ═══════════════════════════════════════════════════════════════════════
def extract_yearly_financials(soup):
    """
    Extract yearly financial table rows.
    Returns dict like:
      {"Revenue": {"FY24": "215", "FY23": "180"}, "PAT": {...}, ...}
    Limited to last 3 year columns.
    """
    financials = {}
    tables = soup.find_all("table")

    for table in tables:
        txt = table.get_text(" ", strip=True).lower()
        # Match financial tables — they contain metrics like revenue, profit, etc.
        if any(kw in txt for kw in ["revenue", "pat", "profit", "ebitda", "net worth",
                                      "eps", "total income", "total assets", "cash flow",
                                      "net income", "equity", "debt", "borrowing"]):
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue

            # Headers = year labels
            headers = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]

            # Determine which columns are the last 3 years
            year_indices = []  # (col_index, year_label)
            for ci, h in enumerate(headers):
                if ci == 0:
                    continue  # skip metric name column
                if h:
                    year_indices.append((ci, h))

            # Take last 3 year columns only
            year_indices = year_indices[-3:] if len(year_indices) > 3 else year_indices

            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all(["th", "td"])]
                if len(cells) < 2 or not cells[0].strip():
                    continue
                metric = cells[0].strip()
                metric_data = {}
                for ci, year_label in year_indices:
                    if ci < len(cells):
                        metric_data[year_label] = cv(cells[ci])
                if metric_data:
                    financials[metric] = metric_data
            break  # first matching table only

    return financials


def scrape_detail(url, page):
    """Scrape one detail page completely."""
    data = {"url": url}

    try:
        page.goto(url, wait_until="networkidle", timeout=45000)
        time.sleep(2)
    except Exception as e:
        data["error"] = str(e)
        return data

    soup = BeautifulSoup(page.content(), "html.parser")

    # ── NAME ──
    h1 = soup.find("h1")
    if h1:
        name = h1.get_text(strip=True)
        name = re.sub(r"\s*\[.*?\]\s*$", "", name)
        name = re.sub(r"\s*E\s*unread\s*messages\s*$", "", name, flags=re.I)
        data["name"] = name.strip()

    # ── PRICE + CHANGE ──
    h4 = soup.find("h4")
    if h4:
        t = h4.get_text(" ", strip=True)
        pm = re.search(r"₹\s*([\d,]+(?:\.\d+)?)", t)
        if pm:
            data["price"] = cv(pm.group(1))
        cm = re.search(r"([-+]?\d+(?:\.\d+)?)\s*\(([-+]?\d+(?:\.\d+)?)(?:%|\))", t)
        if cm:
            data["change_abs"] = cm.group(1)
            data["change_pct"] = cm.group(2)

    # ── FUNDAMENTALS ──
    all_text = list(soup.stripped_strings)
    for i, t in enumerate(all_text):
        if t in FUND_KEYS:
            for j in range(i + 1, min(i + 4, len(all_text))):
                cand = all_text[j]
                if cand and cand not in FUND_KEYS and cand not in ["Buy", "Sell", "Per Equity Share"]:
                    col = t.replace("(in cr.)", "Cr").replace("(%)", "Pct").strip()
                    data[col] = cv(cand)
                    break

    # ── FULL DESCRIPTION (click Read more) ──
    try:
        rm = page.locator("text=Read more").first
        if rm.is_visible(timeout=2000):
            rm.click()
            time.sleep(1)
    except Exception:
        pass

    soup2 = BeautifulSoup(page.content(), "html.parser")
    about = soup2.find(string=re.compile(r"^About\s", re.I))
    if about:
        parent = about.find_parent()
        if parent:
            parts = []
            for sib in parent.find_next_siblings():
                txt = sib.get_text(" ", strip=True)
                if not txt:
                    continue
                if any(txt.startswith(s) for s in
                       ["Fundamentals", "Download App", "Financials",
                        "Reach Out", "Frequently Asked"]):
                    break
                parts.append(txt)
            data["description"] = "\n\n".join(parts)  # FULL text

    # ── YEARLY FINANCIALS (click Yearly tab, extract table) ──
    try:
        yearly_btn = page.locator("text=Yearly").first
        if yearly_btn.is_visible(timeout=3000):
            yearly_btn.click()
            time.sleep(2)
            page.wait_for_load_state("networkidle", timeout=10000)
            time.sleep(1)
    except Exception:
        pass

    soup_fin = BeautifulSoup(page.content(), "html.parser")
    financials = extract_yearly_financials(soup_fin)

    # Flatten: "Revenue [FY24]" = "215"
    for metric, year_vals in financials.items():
        for year, val in year_vals.items():
            data[f"{metric} [{year}]"] = val

    return data


def scrape_all(urls, limit=None):
    from playwright.sync_api import sync_playwright

    total = min(len(urls), limit) if limit else len(urls)
    print(f"\n[2/3] Scraping {total} detail pages...")
    print(f"   Est. time: ~{total * 8 // 60} min\n")

    results, failed = [], []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        pg = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        ).new_page()

        for i, url in enumerate(urls[:total]):
            slug = url.split("/shares/")[-1].rstrip("/")
            print(f"   [{i+1}/{total}] {slug[:50]}...", end=" ", flush=True)

            try:
                d = scrape_detail(url, pg)
                results.append(d)
                if "error" in d:
                    print(f"ERR: {d['error'][:50]}")
                    failed.append(url)
                else:
                    nf = sum(1 for k in d if "[" in k)  # financial columns
                    print(f"OK — {nf} financial data points")
            except Exception as e:
                print(f"CRASH: {str(e)[:50]}")
                results.append({"url": url, "error": str(e)})
                failed.append(url)

            if i < total - 1:
                time.sleep(DELAY)

            if (i + 1) % 50 == 0:
                OUTPUT_DIR.mkdir(exist_ok=True)
                with open(OUTPUT_DIR / "checkpoint.json", "w", encoding="utf-8") as f:
                    json.dump(results, f, indent=2, ensure_ascii=False)
                print(f"   [Checkpoint: {len(results)} saved]")

        browser.close()

    if failed:
        print(f"\n   {len(failed)} failures → output/failed_urls.json")
        OUTPUT_DIR.mkdir(exist_ok=True)
        with open(OUTPUT_DIR / "failed_urls.json", "w") as f:
            json.dump(failed, f, indent=2)

    return results


# ═══════════════════════════════════════════════════════════════════════
# STEP 3: EXPORT
# ═══════════════════════════════════════════════════════════════════════
def export_xlsx(shares, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    hfill = PatternFill("solid", fgColor="1F4E79")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    df = Font(size=9, name="Arial")

    def write_sheet(ws, cols, shares_data):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font, c.fill, c.alignment, c.border = hf, hfill, ha, bd
        for ri, s in enumerate(shares_data, 2):
            for ci, col in enumerate(cols, 1):
                val = s.get(col, "")
                # Numeric conversion for non-text fields
                if col not in ("name", "url", "description", "ISIN Number", "CIN",
                               "PAN Number", "RTA", "Depository", "error"):
                    try:
                        val = float(str(val).replace(",", "").replace("%", "")) if val else ""
                    except (ValueError, TypeError):
                        pass
                c = ws.cell(row=ri, column=ci, value=val)
                c.font, c.border = df, bd
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.column_dimensions["A"].width = 42

    # ── Sheet 1: Overview ──
    ws1 = wb.active
    ws1.title = "Overview"
    core = ["name", "price", "change_abs", "change_pct",
            "Lot Size", "52 Week High", "52 Week Low", "Market Cap Cr",
            "P/E Ratio", "P/B Ratio", "Debt to Equity", "ROE Pct",
            "Book Value", "Face Value", "Total Shares",
            "ISIN Number", "CIN", "PAN Number", "RTA", "Depository", "url"]
    write_sheet(ws1, core, shares)
    ws1.column_dimensions[ws1.cell(1, len(core)).column_letter].width = 55  # url col

    # ── Sheet 2: Yearly Financials ──
    ws2 = wb.create_sheet("Yearly Financials")
    fin_cols = sorted(set(k for s in shares for k in s if "[" in k))
    # Group by metric for readability
    fin_cols.sort(key=lambda x: (x.split("[")[0].strip(), x.split("[")[-1] if "[" in x else ""))
    write_sheet(ws2, ["name"] + fin_cols, shares)

    # ── Sheet 3: Descriptions ──
    ws3 = wb.create_sheet("Descriptions")
    ws3.cell(row=1, column=1, value="Company").font = hf
    ws3.cell(row=1, column=1).fill = hfill
    ws3.cell(row=1, column=2, value="Full Description").font = hf
    ws3.cell(row=1, column=2).fill = hfill
    for ri, s in enumerate(shares, 2):
        ws3.cell(row=ri, column=1, value=s.get("name", "")).font = Font(size=9, name="Arial", bold=True)
        desc = s.get("description", "")
        ws3.cell(row=ri, column=2, value=desc[:32000]).font = df
        ws3.cell(row=ri, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["B"].width = 120
    ws3.freeze_panes = "A2"

    wb.save(filename)
    print(f"   Saved: {filename}")
    print(f"   Sheets: Overview | Yearly Financials | Descriptions")


def export_json(shares, filename):
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(shares, f, indent=2, ensure_ascii=False)
    print(f"   Saved: {filename}")


def export_csv(shares, filename):
    if not shares:
        return
    keys = []
    for s in shares:
        for k in s:
            if k not in keys:
                keys.append(k)
    with open(filename, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=keys)
        w.writeheader()
        w.writerows(shares)
    print(f"   Saved: {filename}")


# ═══════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="UnlistedZone Scraper v4")
    parser.add_argument("--list-only", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--output", default="unlistedzone", help="Filename prefix")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    urls = scrape_listing_page()
    if not urls:
        print("No shares found.")
        sys.exit(1)

    if args.list_only:
        export_json([{"url": u} for u in urls], OUTPUT_DIR / f"{args.output}_urls_{ts}.json")
        return

    shares = scrape_all(urls, limit=args.limit)

    print(f"\n[3/3] Exporting {len(shares)} shares...")
    base = OUTPUT_DIR / f"{args.output}_{ts}"
    export_json(shares, f"{base}.json")
    export_csv(shares, f"{base}.csv")
    try:
        export_xlsx(shares, f"{base}.xlsx")
    except ImportError:
        print("   openpyxl missing, skipping xlsx")

    wf = sum(1 for s in shares if any("[" in k for k in s))
    wd = sum(1 for s in shares if s.get("description"))
    print(f"\n✓ {len(shares)} companies | {wf} with financials | {wd} with descriptions")


if __name__ == "__main__":
    main()
