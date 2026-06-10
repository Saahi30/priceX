"""
UnlistedZone.com Scraper v3 — Full Data Edition
=================================================
- Complete financial data: Yearly + Quarterly tables, all years
- Full description text (no truncation)
- Names from <h1> on detail page
- All fundamentals

Requirements:
  pip install playwright beautifulsoup4 openpyxl requests
  playwright install chromium

Usage:
  python unlistedzone_scraper_v3.py              # Full scrape
  python unlistedzone_scraper_v3.py --limit 5    # Test with 5
  python unlistedzone_scraper_v3.py --list-only  # URLs only
"""

import argparse, csv, json, re, sys, time
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup

BASE_URL = "https://unlistedzone.com"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"}
DELAY = 2.5
OUTPUT_DIR = Path("output")

FUND_KEYS = [
    "Lot Size", "52 Week High", "52 Week Low", "Depository",
    "PAN Number", "ISIN Number", "CIN", "RTA",
    "Market Cap (in cr.)", "P/E Ratio", "P/B Ratio",
    "Debt to Equity", "ROE (%)", "Book Value", "Face Value", "Total Shares",
]


def clean_val(v):
    return v.replace("₹", "").replace(",", "").strip()


# ═══════════════════════════════════════════════════════════════════════
# STEP 1: LISTING PAGE
# ═══════════════════════════════════════════════════════════════════════
def scrape_listing_page():
    from playwright.sync_api import sync_playwright

    print("[1/3] Loading listing page...")
    share_urls = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"{BASE_URL}/shares", wait_until="networkidle", timeout=60000)
        time.sleep(3)

        clicks = 0
        while True:
            try:
                btns = page.locator("text=View More")
                found = False
                for i in range(btns.count()):
                    btn = btns.nth(i)
                    href = btn.get_attribute("href") or ""
                    if "javascript:void" in href:
                        if btn.is_visible(timeout=2000):
                            prev = page.locator("a[href*='/shares/']").count()
                            btn.click()
                            time.sleep(2)
                            page.wait_for_load_state("networkidle", timeout=15000)
                            curr = page.locator("a[href*='/shares/']").count()
                            clicks += 1
                            print(f"   View More #{clicks} — {curr} links")
                            if curr == prev:
                                time.sleep(3)
                                if page.locator("a[href*='/shares/']").count() == prev:
                                    break
                            found = True
                            break
                if not found:
                    break
            except Exception:
                break

        html = page.content()
        browser.close()

    soup = BeautifulSoup(html, "html.parser")
    seen = set()
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith("/shares/"):
            full = BASE_URL + href
        elif "unlistedzone.com/shares/" in href:
            full = href.replace("https://www.unlistedzone.com", BASE_URL)
        else:
            continue
        path = full.replace(BASE_URL, "").rstrip("/")
        if path == "/shares" or path == "":
            continue
        if re.match(r"^/shares/[a-z0-9]", path):
            clean = full.rstrip("/") + "/"
            if clean not in seen:
                seen.add(clean)
                share_urls.append(clean)

    print(f"   Found {len(share_urls)} unique share URLs")
    return share_urls


# ═══════════════════════════════════════════════════════════════════════
# STEP 2: DETAIL PAGES — FULL DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════
def extract_table(soup, label=""):
    """Extract a financial table into a list of dicts: [{metric, year1, year2, ...}]"""
    rows_data = []
    tables = soup.find_all("table")
    for table in tables:
        txt = table.get_text(" ", strip=True).lower()
        if any(kw in txt for kw in ["revenue", "pat", "profit", "ebitda", "net worth", "eps", "total income"]):
            rows = table.find_all("tr")
            if len(rows) < 2:
                continue
            # Headers
            hdrs = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]
            for row in rows[1:]:
                cells = [td.get_text(strip=True) for td in row.find_all(["th", "td"])]
                if len(cells) < 2 or not cells[0].strip():
                    continue
                metric = cells[0].strip()
                entry = {"metric": f"{label}{metric}"}
                for ci in range(1, len(cells)):
                    year_label = hdrs[ci] if ci < len(hdrs) else f"Col{ci}"
                    entry[year_label] = clean_val(cells[ci])
                rows_data.append(entry)
            break
    return rows_data


def scrape_detail(url, page):
    """Full detail page scrape: name, price, fundamentals, description, yearly + quarterly financials."""
    data = {"url": url}

    try:
        page.goto(url, wait_until="networkidle", timeout=45000)
        time.sleep(2)
    except Exception as e:
        data["error"] = str(e)
        return data

    html = page.content()
    soup = BeautifulSoup(html, "html.parser")

    # ── NAME from <h1> ──
    h1 = soup.find("h1")
    if h1:
        name = h1.get_text(strip=True)
        name = re.sub(r"\s*\[.*?\]\s*$", "", name)
        name = re.sub(r"\s*E\s*unread\s*messages\s*$", "", name, flags=re.I)
        data["name"] = name.strip()

    # ── PRICE + CHANGE ──
    h4 = soup.find("h4")
    if h4:
        h4t = h4.get_text(" ", strip=True)
        pm = re.search(r"₹\s*([\d,]+(?:\.\d+)?)", h4t)
        if pm:
            data["price"] = clean_val(pm.group(1))
        cm = re.search(r"([-+]?\d+(?:\.\d+)?)\s*\(([-+]?\d+(?:\.\d+)?)(?:%|\))", h4t)
        if cm:
            data["change_abs"] = cm.group(1)
            data["change_pct"] = cm.group(2)

    # ── FUNDAMENTALS ──
    all_text = list(soup.stripped_strings)
    for i, t in enumerate(all_text):
        if t in FUND_KEYS:
            for j in range(i + 1, min(i + 4, len(all_text))):
                cand = all_text[j]
                if cand and cand not in FUND_KEYS and cand not in ["Buy", "Sell", "Per Equity Share"]:
                    col = t.replace("(in cr.)", "Cr").replace("(%)", "Pct").strip()
                    data[col] = clean_val(cand)
                    break

    # ── FULL DESCRIPTION (no truncation) ──
    about_header = soup.find(string=re.compile(r"^About\s", re.I))
    if about_header:
        parent = about_header.find_parent()
        if parent:
            # First try: look for a "Read more" expandable section
            # Click "Read more" if present
            try:
                read_more = page.locator("text=Read more").first
                if read_more.is_visible(timeout=2000):
                    read_more.click()
                    time.sleep(1)
                    # Re-parse after expanding
                    html = page.content()
                    soup = BeautifulSoup(html, "html.parser")
                    about_header = soup.find(string=re.compile(r"^About\s", re.I))
                    parent = about_header.find_parent() if about_header else None
            except Exception:
                pass

            if parent:
                desc_parts = []
                for sib in parent.find_next_siblings():
                    text = sib.get_text(" ", strip=True)
                    if not text:
                        continue
                    # Stop at next major section
                    if any(text.startswith(s) for s in ["Fundamentals", "Download App", "Financials", "Reach Out", "Frequently Asked"]):
                        break
                    desc_parts.append(text)
                data["description"] = "\n\n".join(desc_parts)  # Full text, no truncation

    # ── YEARLY FINANCIALS ──
    try:
        yearly_btn = page.locator("text=Yearly").first
        if yearly_btn.is_visible(timeout=3000):
            yearly_btn.click()
            time.sleep(2)
            page.wait_for_load_state("networkidle", timeout=10000)
            time.sleep(1)
    except Exception:
        pass

    html_y = page.content()
    soup_y = BeautifulSoup(html_y, "html.parser")
    yearly_rows = extract_table(soup_y, label="Y: ")

    # Flatten yearly into data
    for row in yearly_rows:
        metric = row.pop("metric")
        for year, val in row.items():
            data[f"{metric} [{year}]"] = val

    # ── QUARTERLY FINANCIALS ──
    try:
        quarterly_btn = page.locator("text=Quarterly").first
        if quarterly_btn.is_visible(timeout=3000):
            quarterly_btn.click()
            time.sleep(2)
            page.wait_for_load_state("networkidle", timeout=10000)
            time.sleep(1)
    except Exception:
        pass

    html_q = page.content()
    soup_q = BeautifulSoup(html_q, "html.parser")
    quarterly_rows = extract_table(soup_q, label="Q: ")

    for row in quarterly_rows:
        metric = row.pop("metric")
        for qtr, val in row.items():
            data[f"{metric} [{qtr}]"] = val

    return data


def scrape_all_details(share_urls, limit=None):
    from playwright.sync_api import sync_playwright

    total = len(share_urls) if limit is None else min(limit, len(share_urls))
    print(f"\n[2/3] Scraping {total} detail pages (full financials + descriptions)...")
    print(f"   Estimated time: ~{total * 8 // 60} min\n")

    results = []
    failed = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(
            viewport={"width": 1280, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
        )
        pg = ctx.new_page()

        for i, url in enumerate(share_urls[:total]):
            slug = url.split("/shares/")[-1].rstrip("/")
            print(f"   [{i+1}/{total}] {slug[:50]}...", end=" ", flush=True)

            try:
                data = scrape_detail(url, pg)
                results.append(data)
                if "error" in data:
                    print(f"ERROR: {data['error'][:60]}")
                    failed.append(url)
                else:
                    n_fields = len([k for k in data if k not in ["url", "error"]])
                    has_fin = any(k.startswith("Y:") or k.startswith("Q:") for k in data)
                    fin_tag = " +financials" if has_fin else ""
                    print(f"OK ({n_fields} fields{fin_tag})")
            except Exception as e:
                print(f"CRASH: {str(e)[:60]}")
                results.append({"url": url, "error": str(e)})
                failed.append(url)

            if i < total - 1:
                time.sleep(DELAY)

            # Save checkpoint every 50
            if (i + 1) % 50 == 0:
                ckpt = OUTPUT_DIR / "checkpoint.json"
                with open(ckpt, "w", encoding="utf-8") as f:
                    json.dump(results, f, indent=2, ensure_ascii=False)
                print(f"   [Checkpoint saved: {len(results)} shares]")

        browser.close()

    if failed:
        print(f"\n   {len(failed)} pages failed. URLs saved to output/failed_urls.json")
        with open(OUTPUT_DIR / "failed_urls.json", "w") as f:
            json.dump(failed, f, indent=2)

    return results


# ═══════════════════════════════════════════════════════════════════════
# STEP 3: EXPORT
# ═══════════════════════════════════════════════════════════════════════
def export_xlsx(shares, filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # ── SHEET 1: OVERVIEW (fundamentals + latest financials) ──
    ws1 = wb.active
    ws1.title = "Overview"

    core_cols = ["name", "price", "change_abs", "change_pct",
                 "Lot Size", "52 Week High", "52 Week Low", "Market Cap Cr",
                 "P/E Ratio", "P/B Ratio", "Debt to Equity", "ROE Pct",
                 "Book Value", "Face Value", "Total Shares",
                 "ISIN Number", "CIN", "PAN Number", "RTA", "Depository", "url"]

    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    for ci, col in enumerate(core_cols, 1):
        cell = ws1.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, border

    for ri, share in enumerate(shares, 2):
        for ci, col in enumerate(core_cols, 1):
            val = share.get(col, "")
            if col not in ["name", "url", "ISIN Number", "CIN", "PAN Number", "RTA", "Depository"]:
                try:
                    val = float(str(val).replace(",", "").replace("%", "")) if val else ""
                except (ValueError, TypeError):
                    pass
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9, name="Arial")
            cell.border = border

    for ci, col in enumerate(core_cols, 1):
        w = 14
        if col == "name": w = 42
        elif col == "url": w = 55
        ws1.column_dimensions[ws1.cell(1, ci).column_letter].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── SHEET 2: YEARLY FINANCIALS ──
    ws2 = wb.create_sheet("Yearly Financials")
    y_cols = ["name"]
    for s in shares:
        for k in s:
            if k.startswith("Y: ") and k not in y_cols:
                y_cols.append(k)
    y_cols.sort(key=lambda x: (x.split("[")[-1] if "[" in x else x) if x != "name" else "")
    y_cols = ["name"] + [c for c in y_cols if c != "name"]

    for ci, col in enumerate(y_cols, 1):
        cell = ws2.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, border

    for ri, share in enumerate(shares, 2):
        for ci, col in enumerate(y_cols, 1):
            val = share.get(col, "")
            if col != "name":
                try:
                    val = float(str(val).replace(",", "").replace("%", "")) if val else ""
                except (ValueError, TypeError):
                    pass
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9, name="Arial")
            cell.border = border

    if y_cols:
        ws2.column_dimensions["A"].width = 42
        ws2.freeze_panes = "B2"

    # ── SHEET 3: QUARTERLY FINANCIALS ──
    ws3 = wb.create_sheet("Quarterly Financials")
    q_cols = ["name"]
    for s in shares:
        for k in s:
            if k.startswith("Q: ") and k not in q_cols:
                q_cols.append(k)
    q_cols.sort(key=lambda x: (x.split("[")[-1] if "[" in x else x) if x != "name" else "")
    q_cols = ["name"] + [c for c in q_cols if c != "name"]

    for ci, col in enumerate(q_cols, 1):
        cell = ws3.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, border

    for ri, share in enumerate(shares, 2):
        for ci, col in enumerate(q_cols, 1):
            val = share.get(col, "")
            if col != "name":
                try:
                    val = float(str(val).replace(",", "").replace("%", "")) if val else ""
                except (ValueError, TypeError):
                    pass
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9, name="Arial")
            cell.border = border

    if q_cols:
        ws3.column_dimensions["A"].width = 42
        ws3.freeze_panes = "B2"

    # ── SHEET 4: DESCRIPTIONS ──
    ws4 = wb.create_sheet("Descriptions")
    ws4.cell(row=1, column=1, value="Company").font = hdr_font
    ws4.cell(row=1, column=1).fill = hdr_fill
    ws4.cell(row=1, column=2, value="Description").font = hdr_font
    ws4.cell(row=1, column=2).fill = hdr_fill

    for ri, share in enumerate(shares, 2):
        ws4.cell(row=ri, column=1, value=share.get("name", "")).font = Font(size=9, name="Arial", bold=True)
        desc = share.get("description", "")
        ws4.cell(row=ri, column=2, value=desc[:32000]).font = Font(size=9, name="Arial")  # Excel cell limit ~32k
        ws4.cell(row=ri, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 120
    ws4.freeze_panes = "A2"

    wb.save(filename)
    print(f"   Saved: {filename}")
    print(f"   Sheets: Overview | Yearly Financials | Quarterly Financials | Descriptions")


def export_csv(shares, filename):
    if not shares:
        return
    all_keys = []
    for s in shares:
        for k in s:
            if k not in all_keys:
                all_keys.append(k)
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys)
        writer.writeheader()
        writer.writerows(shares)
    print(f"   Saved: {filename}")


def export_json(shares, filename):
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(shares, f, indent=2, ensure_ascii=False)
    print(f"   Saved: {filename}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="UnlistedZone Scraper v3 — Full Data")
    parser.add_argument("--list-only", action="store_true")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--output", default="unlistedzone_full", help="Output filename prefix")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")

    share_urls = scrape_listing_page()
    if not share_urls:
        print("ERROR: No shares found.")
        sys.exit(1)

    if args.list_only:
        data = [{"url": u} for u in share_urls]
        export_json(data, OUTPUT_DIR / f"{args.output}_urls_{ts}.json")
        print(f"\nDone! {len(share_urls)} URLs.")
        return

    shares = scrape_all_details(share_urls, limit=args.limit)

    print(f"\n[3/3] Exporting {len(shares)} shares...")
    base = OUTPUT_DIR / f"{args.output}_{ts}"
    export_csv(shares, f"{base}.csv")
    export_json(shares, f"{base}.json")
    try:
        export_xlsx(shares, f"{base}.xlsx")
    except ImportError:
        print("   openpyxl not installed, skipping xlsx")

    # Stats
    with_fin = sum(1 for s in shares if any(k.startswith("Y:") for k in s))
    with_desc = sum(1 for s in shares if s.get("description", ""))
    print(f"\n✓ {len(shares)} companies scraped")
    print(f"  {with_fin} with yearly financials")
    print(f"  {with_desc} with descriptions")


if __name__ == "__main__":
    main()
