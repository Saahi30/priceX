"""
UnlistedZone.com Full Scraper
==============================
Scrapes all unlisted shares from unlistedzone.com:
  1. Listing page: company name, URL, current price, change
  2. Detail pages: fundamentals (P/E, P/B, Market Cap, ROE, etc.)

Requirements:
  pip install playwright beautifulsoup4 openpyxl requests
  playwright install chromium

Usage:
  python unlistedzone_scraper.py              # Full scrape (listing + details)
  python unlistedzone_scraper.py --list-only  # Only scrape listing page
  python unlistedzone_scraper.py --limit 10   # Scrape only first 10 detail pages
"""

import argparse
import csv
import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup

# ─── CONFIG ───────────────────────────────────────────────────────────
BASE_URL = "https://unlistedzone.com"
LISTING_URL = f"{BASE_URL}/shares"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
}
DELAY_BETWEEN_REQUESTS = 1.5  # seconds between detail page requests (be respectful)
OUTPUT_DIR = Path("output")


# ─── STEP 1: SCRAPE LISTING PAGE (Playwright for JS rendering) ───────
def scrape_listing_page():
    """Use Playwright to load all shares (clicks 'View More' until exhausted)."""
    from playwright.sync_api import sync_playwright

    print("[1/3] Launching browser to scrape listing page...")
    shares = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(LISTING_URL, wait_until="networkidle", timeout=60000)
        time.sleep(3)

        # Click "View More" until it disappears or stops loading new content
        click_count = 0
        while True:
            try:
                view_more = page.locator("a:has-text('View More')").first
                if not view_more.is_visible(timeout=3000):
                    break
                prev_count = page.locator("a[href*='/shares/']").count()
                view_more.click()
                time.sleep(2)
                page.wait_for_load_state("networkidle", timeout=15000)
                new_count = page.locator("a[href*='/shares/']").count()
                click_count += 1
                print(f"   Clicked 'View More' #{click_count} — {new_count} links found")
                if new_count == prev_count:
                    # Try one more time
                    time.sleep(3)
                    final_count = page.locator("a[href*='/shares/']").count()
                    if final_count == prev_count:
                        break
            except Exception:
                break

        # Parse the fully loaded page
        html = page.content()
        browser.close()

    soup = BeautifulSoup(html, "html.parser")

    # Extract share cards — each card has a link and price info
    # The structure: each share is in a card with an <a> containing the company name/URL
    # and a price displayed nearby
    seen_urls = set()
    all_links = soup.find_all("a", href=re.compile(r"/shares/[^/]+/?$"))

    for link in all_links:
        href = link.get("href", "")
        if href in seen_urls or href == "/shares" or href == "/shares/":
            continue

        full_url = href if href.startswith("http") else BASE_URL + href
        name = link.get_text(strip=True)

        if not name or len(name) < 3:
            continue

        # Try to find price near this link
        parent = link.find_parent()
        card = parent
        for _ in range(5):
            if card is None:
                break
            card_text = card.get_text(" ", strip=True)
            if "₹" in card_text:
                break
            card = card.find_parent()

        price = ""
        change_abs = ""
        change_pct = ""
        if card:
            card_text = card.get_text(" ", strip=True)
            # Extract price: ₹XXX
            price_match = re.search(r"₹\s*([\d,]+(?:\.\d+)?)", card_text)
            if price_match:
                price = price_match.group(1).replace(",", "")
            # Extract change: (+/-X.XX) (X.XX%)
            change_match = re.search(r"\(([-+]?\d+(?:\.\d+)?)\s*\)\s*\(([-+]?\d+(?:\.\d+)?)%\)", card_text)
            if change_match:
                change_abs = change_match.group(1)
                change_pct = change_match.group(2)

        seen_urls.add(href)
        shares.append({
            "name": name,
            "url": full_url,
            "price": price,
            "change_abs": change_abs,
            "change_pct": change_pct,
        })

    print(f"   Found {len(shares)} unique shares on listing page")
    return shares


# ─── STEP 2: SCRAPE DETAIL PAGES ─────────────────────────────────────
FUNDAMENTALS_KEYS = [
    "Lot Size", "52 Week High", "52 Week Low", "Depository",
    "PAN Number", "ISIN Number", "CIN", "RTA",
    "Market Cap (in cr.)", "P/E Ratio", "P/B Ratio",
    "Debt to Equity", "ROE (%)", "Book Value", "Face Value", "Total Shares",
]


def scrape_detail_page(url, session):
    """Scrape fundamentals from a single share detail page."""
    try:
        r = session.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        return {"error": str(e)}

    soup = BeautifulSoup(r.text, "html.parser")
    data = {}

    # Extract the about/description section
    about_section = soup.find(string=re.compile(r"About\s", re.I))
    if about_section:
        parent = about_section.find_parent()
        if parent:
            next_sib = parent.find_next_sibling()
            if next_sib:
                data["description"] = next_sib.get_text(" ", strip=True)[:500]

    # Extract fundamentals - they appear as key-value pairs in the page
    page_text = soup.get_text(" ", strip=True)

    for key in FUNDAMENTALS_KEYS:
        # Look for the key in text and grab the value after it
        pattern = re.escape(key) + r"\s*[:\s]*₹?\s*([\d,]+(?:\.\d+)?%?)"
        match = re.search(pattern, page_text)
        if match:
            val = match.group(1).replace(",", "")
            col_name = key.replace("(in cr.)", "Cr").replace("(%)", "Pct").strip()
            data[col_name] = val

    # Also try to extract ISIN, CIN, PAN, RTA as text (non-numeric)
    for key in ["ISIN Number", "CIN", "PAN Number", "RTA", "Depository"]:
        pattern = re.escape(key) + r"\s*[:\s]*([\w\d&\s]+?)(?:\s{2,}|$)"
        match = re.search(pattern, page_text)
        if match:
            col_name = key.strip()
            data[col_name] = match.group(1).strip()

    # Try extracting from structured HTML elements more precisely
    # Look for table-like structures or definition lists
    all_text_elements = soup.find_all(string=True)
    for i, el in enumerate(all_text_elements):
        text = el.strip()
        if text in FUNDAMENTALS_KEYS:
            # Get the next non-empty text element
            for j in range(i + 1, min(i + 5, len(all_text_elements))):
                val = all_text_elements[j].strip()
                if val and val != text:
                    col_name = text.replace("(in cr.)", "Cr").replace("(%)", "Pct").strip()
                    data[col_name] = val.replace("₹", "").replace(",", "").strip()
                    break

    return data


def scrape_all_details(shares, limit=None):
    """Scrape detail pages for all shares."""
    total = len(shares) if limit is None else min(limit, len(shares))
    print(f"\n[2/3] Scraping {total} detail pages...")

    session = requests.Session()

    for i, share in enumerate(shares[:total]):
        url = share["url"]
        print(f"   [{i+1}/{total}] {share['name'][:50]}...", end=" ", flush=True)

        details = scrape_detail_page(url, session)
        share.update(details)

        if "error" in details:
            print(f"ERROR: {details['error']}")
        else:
            print(f"OK ({len(details)} fields)")

        if i < total - 1:
            time.sleep(DELAY_BETWEEN_REQUESTS)

    return shares


# ─── STEP 3: EXPORT ──────────────────────────────────────────────────
def export_to_xlsx(shares, filename):
    """Export to a formatted Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Unlisted Shares"

    # Determine all columns
    base_cols = ["name", "url", "price", "change_abs", "change_pct"]
    detail_cols = []
    for s in shares:
        for k in s:
            if k not in base_cols and k not in detail_cols and k != "error":
                detail_cols.append(k)

    all_cols = base_cols + sorted(detail_cols)

    # Header formatting
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Pretty column names
    col_display = {
        "name": "Company Name", "url": "URL", "price": "Price (₹)",
        "change_abs": "Change (₹)", "change_pct": "Change (%)",
        "description": "Description", "Lot Size": "Lot Size",
        "52 Week High": "52W High", "52 Week Low": "52W Low",
        "Market Cap Cr": "MCap (Cr)", "P/E Ratio": "P/E",
        "P/B Ratio": "P/B", "Debt to Equity": "D/E",
        "ROE Pct": "ROE %", "Book Value": "Book Value",
        "Face Value": "Face Value", "Total Shares": "Total Shares",
        "ISIN Number": "ISIN", "CIN": "CIN", "PAN Number": "PAN",
        "RTA": "RTA", "Depository": "Depository",
    }

    # Write headers
    for col_idx, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_display.get(col, col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    data_font = Font(size=10, name="Arial")
    for row_idx, share in enumerate(shares, 2):
        for col_idx, col in enumerate(all_cols, 1):
            val = share.get(col, "")
            # Try to convert numeric strings
            if col in ["price", "change_abs", "change_pct", "Lot Size", "52 Week High",
                        "52 Week Low", "Market Cap Cr", "P/E Ratio", "P/B Ratio",
                        "Debt to Equity", "ROE Pct", "Book Value", "Face Value", "Total Shares"]:
                try:
                    val = float(str(val).replace(",", "").replace("%", ""))
                except (ValueError, TypeError):
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

    # Auto-width columns
    for col_idx, col in enumerate(all_cols, 1):
        max_len = len(col_display.get(col, col))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"   Saved: {filename}")


def export_to_csv(shares, filename):
    """Export to CSV as fallback."""
    if not shares:
        return
    all_keys = []
    for s in shares:
        for k in s:
            if k not in all_keys:
                all_keys.append(k)

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys)
        writer.writeheader()
        writer.writerows(shares)
    print(f"   Saved: {filename}")


def export_to_json(shares, filename):
    """Export to JSON."""
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(shares, f, indent=2, ensure_ascii=False)
    print(f"   Saved: {filename}")


# ─── MAIN ─────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Scrape UnlistedZone.com shares data")
    parser.add_argument("--list-only", action="store_true", help="Only scrape listing page")
    parser.add_argument("--limit", type=int, default=None, help="Limit detail pages to scrape")
    parser.add_argument("--output", default="unlistedzone_shares", help="Output filename (no ext)")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    # Step 1: Listing page
    shares = scrape_listing_page()
    if not shares:
        print("ERROR: No shares found. The page structure may have changed.")
        sys.exit(1)

    # Step 2: Detail pages (optional)
    if not args.list_only:
        shares = scrape_all_details(shares, limit=args.limit)

    # Step 3: Export
    print(f"\n[3/3] Exporting {len(shares)} shares...")
    base = OUTPUT_DIR / f"{args.output}_{timestamp}"

    export_to_csv(shares, f"{base}.csv")
    export_to_json(shares, f"{base}.json")
    try:
        export_to_xlsx(shares, f"{base}.xlsx")
    except ImportError:
        print("   (openpyxl not installed, skipping xlsx)")

    print(f"\nDone! {len(shares)} shares scraped.")


if __name__ == "__main__":
    main()
